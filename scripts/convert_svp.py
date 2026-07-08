#!/usr/bin/env python3
"""Convert the in-progress study sheets of Swedish-Kelly.xlsx into JSON.

The "Swedish in-progress" mode of the PWA has two tabs, each backed by one
sheet of the Kelly workbook:

    Learn today  -> svp/learn-today.json   (tab "Learn Today", the daily batch)
    Yellow       -> svp/yellow.json        (tab "Yello", capped at MAX_ROW)

Sheet columns (row 1 is a header):
    Index, Article, Swedish Word, English Meaning,
    Part of Speech, Swedish Sentence, English Sentence

The article (att / en / ett) is merged into the displayed word, e.g.
"vara" + "att" -> "att vara", "väg" + "en" -> "en väg" — matching how
swedish.json is built (see convert_swedish.py).

Each row's word cell (col C) is filled with its status color. We read that
fill and tag red-status words so the PWA can show a red dot next to them:
    Red    FFC7CE  -> "status": "red"     (don't know / slipped back)
    Yellow FFF3A0  -> (default, no status field)

Output entry shape matches swedish.json so the app renders every dataset
identically (`fi`/`fiSentence` hold the *target* Swedish word/sentence):
    { index, fi, en, fiSentence, enSentence, pos, hasAudio, status? }

`hasAudio` is left False here; generate_audio.py flips it to True once the
clips exist. Both tabs share one audio pool (audio/svp) keyed by index,
because every row comes from the same Sheet2 — so a given index has identical
word/meaning/sentence text wherever it appears.
"""

import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT.parent / "Swedish-Kelly.xlsx"   # workbook lives next to the PWA folder
OUT_DIR = ROOT / "svp"

# sheet name -> (output filename, last row to read incl. header; None = all rows)
SHEETS = {
    "Learn today": ("learn-today.json", None),
    "Yellow": ("yellow.json", 142),
}

RED_FILL = "FFC7CE"  # Excel "light red" fill = red status (don't-know)


def cell_status(cell):
    """Return 'red' if the cell's fill is the red-status color, else None."""
    f = cell.fill
    if f and f.patternType == "solid" and f.fgColor and f.fgColor.rgb:
        rgb = str(f.fgColor.rgb)
        if rgb.endswith(RED_FILL):
            return "red"
    return None


def rows_from_sheet(ws, max_row=None):
    entries = []
    last = max_row or ws.max_row
    for r in range(2, last + 1):  # row 1 is the header
        index = ws.cell(r, 1).value
        word = ws.cell(r, 3).value
        if index is None or not word:
            continue
        cells = [ws.cell(r, c).value for c in range(1, 8)]
        index, article, word, meaning, pos, sv_sentence, en_sentence = (
            ("" if c is None else str(c).strip()) for c in cells
        )
        display = f"{article} {word}".strip() if article else word
        entry = {
            "index": index,
            "fi": display,
            "en": meaning,
            "fiSentence": sv_sentence,
            "enSentence": en_sentence,
            "pos": pos,
            "hasAudio": False,
        }
        status = cell_status(ws.cell(r, 3))  # word cell holds the status color
        if status:
            entry["status"] = status
        entries.append(entry)
    return entries


def main() -> int:
    wb = openpyxl.load_workbook(XLSX, data_only=True)  # not read_only: need fills
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    for sheet, (filename, max_row) in SHEETS.items():
        out = OUT_DIR / filename
        if sheet not in wb.sheetnames:
            print(f"! sheet {sheet!r} not found -> writing empty {filename}")
            out.write_text("[]", encoding="utf-8")
            continue
        entries = rows_from_sheet(wb[sheet], max_row)
        reds = sum(1 for e in entries if e.get("status") == "red")
        out.write_text(
            json.dumps(entries, ensure_ascii=False, indent=0), encoding="utf-8"
        )
        print(f"Wrote {len(entries):>3} entries ({reds} red)  {sheet!r:>13} "
              f"-> {out.relative_to(ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
